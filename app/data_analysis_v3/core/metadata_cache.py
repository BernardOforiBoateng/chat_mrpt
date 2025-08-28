"""
Metadata Cache for Data Analysis V3
Efficiently handles metadata extraction and caching for large files
"""

import os
import json
import logging
import pandas as pd
from .encoding_handler import EncodingHandler
from typing import Dict, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)


class MetadataCache:
    """
    Handles metadata caching for uploaded files.
    Prevents loading entire large files just to get basic info.
    """
    
    # Sampling thresholds
    SMALL_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    SAMPLE_ROWS = 1000  # Sample size for large files
    
    @staticmethod
    def get_cache_path(session_id: str) -> str:
        """Get the path to the metadata cache file."""
        return f"instance/uploads/{session_id}/metadata_cache.json"
    
    @staticmethod
    def load_cache(session_id: str) -> Optional[Dict[str, Any]]:
        """Load cached metadata if it exists."""
        cache_path = MetadataCache.get_cache_path(session_id)
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'r') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Error loading metadata cache: {e}")
        return None
    
    @staticmethod
    def save_cache(session_id: str, metadata: Dict[str, Any]):
        """Save metadata to cache."""
        cache_path = MetadataCache.get_cache_path(session_id)
        try:
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            with open(cache_path, 'w') as f:
                json.dump(metadata, f, indent=2, default=str)
            logger.info(f"Metadata cache saved for session {session_id}")
        except Exception as e:
            logger.error(f"Error saving metadata cache: {e}")
    
    @staticmethod
    def extract_file_metadata(filepath: str, filename: str) -> Dict[str, Any]:
        """
        Extract metadata from a file without loading it entirely if large.
        
        Args:
            filepath: Full path to the file
            filename: Original filename
            
        Returns:
            Dictionary containing file metadata
        """
        metadata = {
            'filename': filename,
            'filepath': filepath,
            'file_size': os.path.getsize(filepath),
            'file_size_mb': round(os.path.getsize(filepath) / (1024 * 1024), 2),
            'upload_time': datetime.now().isoformat(),
            'is_sampled': False
        }
        
        try:
            if filename.endswith('.csv'):
                metadata.update(MetadataCache._extract_csv_metadata(filepath))
            elif filename.endswith(('.xlsx', '.xls')):
                metadata.update(MetadataCache._extract_excel_metadata(filepath))
            elif filename.endswith('.json'):
                metadata.update(MetadataCache._extract_json_metadata(filepath))
            else:
                metadata['type'] = 'text'
                metadata['rows'] = 'N/A'
                metadata['columns'] = 'N/A'
        except Exception as e:
            logger.error(f"Error extracting metadata from {filename}: {e}")
            metadata['error'] = str(e)
            metadata['rows'] = 'Error'
            metadata['columns'] = 'Error'
        
        return metadata
    
    @staticmethod
    def _extract_csv_metadata(filepath: str) -> Dict[str, Any]:
        """Extract metadata from CSV file."""
        file_size = os.path.getsize(filepath)
        
        if file_size <= MetadataCache.SMALL_FILE_SIZE:
            # Small file: load fully
            # Check if this is TPR data to decide on sanitization
            df_test = pd.read_csv(filepath, nrows=5)
            is_tpr_data = any('RDT' in col or 'Microscopy' in col or 'TPR' in col for col in df_test.columns)
            
            df = EncodingHandler.read_csv_with_encoding(filepath)
            return {
                'type': 'csv',
                'rows': df.shape[0],
                'columns': df.shape[1],
                'column_names': list(df.columns),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'is_sampled': False
            }
        else:
            # Large file: sample first rows
            # Check if this is TPR data to decide on sanitization
            df_test = pd.read_csv(filepath, nrows=5)
            is_tpr_data = any('RDT' in col or 'Microscopy' in col or 'TPR' in col for col in df_test.columns)
            
            df_sample = EncodingHandler.read_csv_with_encoding(filepath, nrows=MetadataCache.SAMPLE_ROWS)
            
            # Estimate total rows from file size
            sample_size = df_sample.memory_usage(deep=True).sum()
            estimated_rows = int((file_size / sample_size) * MetadataCache.SAMPLE_ROWS)
            
            return {
                'type': 'csv',
                'rows': estimated_rows,
                'rows_estimated': True,
                'columns': df_sample.shape[1],
                'column_names': list(df_sample.columns),
                'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                'is_sampled': True,
                'sample_rows': MetadataCache.SAMPLE_ROWS
            }
    
    @staticmethod
    def _extract_excel_metadata(filepath: str) -> Dict[str, Any]:
        """Extract metadata from Excel file."""
        file_size = os.path.getsize(filepath)
        
        if file_size <= MetadataCache.SMALL_FILE_SIZE:
            # Small file: load fully
            df = pd.read_excel(filepath)
            return {
                'type': 'excel',
                'rows': df.shape[0],
                'columns': df.shape[1],
                'column_names': list(df.columns),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'is_sampled': False
            }
        else:
            # Large file: sample first rows
            # For Excel, we can use nrows parameter
            df_sample = pd.read_excel(filepath, nrows=MetadataCache.SAMPLE_ROWS)
            
            # For Excel files, we can try to get exact row count without loading all data
            # Using openpyxl to get dimensions
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                total_rows = ws.max_row
                wb.close()
                
                return {
                    'type': 'excel',
                    'rows': total_rows,
                    'rows_exact': True,
                    'columns': df_sample.shape[1],
                    'column_names': list(df_sample.columns),
                    'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                    'is_sampled': True,
                    'sample_rows': MetadataCache.SAMPLE_ROWS
                }
            except:
                # Fallback to estimation
                sample_size = df_sample.memory_usage(deep=True).sum()
                # Excel files have overhead, adjust estimation
                estimated_rows = int((file_size / sample_size) * MetadataCache.SAMPLE_ROWS * 0.7)
                
                return {
                    'type': 'excel',
                    'rows': estimated_rows,
                    'rows_estimated': True,
                    'columns': df_sample.shape[1],
                    'column_names': list(df_sample.columns),
                    'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                    'is_sampled': True,
                    'sample_rows': MetadataCache.SAMPLE_ROWS
                }
    
    @staticmethod
    def _extract_json_metadata(filepath: str) -> Dict[str, Any]:
        """Extract metadata from JSON file."""
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        if isinstance(data, list):
            return {
                'type': 'json',
                'rows': len(data),
                'structure': 'array',
                'is_sampled': False
            }
        elif isinstance(data, dict):
            return {
                'type': 'json',
                'keys': list(data.keys()),
                'structure': 'object',
                'is_sampled': False
            }
        else:
            return {
                'type': 'json',
                'structure': type(data).__name__,
                'is_sampled': False
            }
    
    @staticmethod
    def update_file_metadata(session_id: str, filepath: str, filename: str):
        """
        Update metadata cache for a specific file.
        
        Args:
            session_id: User session ID
            filepath: Full path to the file
            filename: Original filename
        """
        # Load existing cache or create new
        cache = MetadataCache.load_cache(session_id) or {'files': {}}
        
        # Extract metadata for this file
        metadata = MetadataCache.extract_file_metadata(filepath, filename)
        
        # Update cache
        cache['files'][filename] = metadata
        cache['last_updated'] = datetime.now().isoformat()
        
        # Save cache
        MetadataCache.save_cache(session_id, cache)
        
        return metadata
    
    @staticmethod
    def get_summary_from_cache(session_id: str) -> Optional[str]:
        """
        Get a formatted summary of available data from cache.
        
        Args:
            session_id: User session ID
            
        Returns:
            Formatted string summary or None if no cache
        """
        cache = MetadataCache.load_cache(session_id)
        if not cache or 'files' not in cache:
            return None
        
        summary_parts = ["Available data:"]
        
        for filename, metadata in cache['files'].items():
            if metadata.get('error'):
                summary_parts.append(f"- {filename}: Error loading file")
            elif metadata.get('type') in ['csv', 'excel'] and 'rows' in metadata:
                rows = metadata.get('rows', 'Unknown')
                cols = metadata.get('columns', 'Unknown')
                
                # Format row count
                if isinstance(rows, int):
                    if metadata.get('rows_estimated'):
                        rows_str = f"~{rows:,}"
                    else:
                        rows_str = f"{rows:,}"
                else:
                    rows_str = str(rows)
                
                # Add size info for large files
                size_mb = metadata.get('file_size_mb', 0)
                if size_mb > 10:
                    size_info = f" ({size_mb:.1f}MB)"
                else:
                    size_info = ""
                
                var_name = filename.split('.')[0].replace(' ', '_').replace('-', '_')
                summary_parts.append(f"- {var_name}: {rows_str} rows, {cols} columns{size_info}")
                
                # Add sampling note if applicable
                if metadata.get('is_sampled'):
                    summary_parts.append(f"  (Metadata from first {metadata.get('sample_rows', 1000)} rows)")
            elif metadata.get('type') == 'json':
                structure = metadata.get('structure', 'Unknown')
                if structure == 'array':
                    summary_parts.append(f"- {filename}: JSON array with {metadata.get('rows', 'Unknown')} items")
                else:
                    summary_parts.append(f"- {filename}: JSON {structure}")
            else:
                summary_parts.append(f"- {filename}: {metadata.get('type', 'Unknown')} file")
        
        return "\n".join(summary_parts)